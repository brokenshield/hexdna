# encoding: utf-8
__version__ = "2.5.00"
__author__ = "Gunnar Roxen <gunnar@brokenshield.net>"

from BSUtilities import BrokenShieldUtilities
from openpyxl import load_workbook
from character_dataclasses import (
    BSCMConfig,
    LiveCharacterModel,
)
import sqlite3
import json
import logging


logging.basicConfig(level=logging.WARNING)


class ExcelImport:
    """This script reads gamedata.xlsx and exports ut to gamedata.json"""

    def __init__(self, **kwargs):
        super(ExcelImport, self).__init__(**kwargs)
        # Create Gamedata DB
        self.create_gamedata_db()

    @staticmethod
    def split_effects(data, want_list=False):
        """Processes cell data and divides up the individual effects for formatting"""
        # TODO: Fix bug where it cannot process a text string that begins with a minus symbol (-)
        bscm = BSCMConfig()
        list_of_effects = {}

        # Check data is not null
        if data:
            # Check if there is more than one entry separated by a comma. If there is split it up
            if ";" in data:
                effects = data.split(";")
                if want_list is False:
                    for effect in effects:
                        individual_effect = effect.split(":")
                        effect_id = individual_effect[0]
                        if effect_id in bscm.convert_to_list:
                            effect_value = individual_effect[1].split(",")
                        else:
                            if individual_effect[1].isnumeric():
                                effect_value = int(individual_effect[1])
                            else:
                                temp_effect_value = individual_effect[1]
                                if temp_effect_value[0] == "-":
                                    effect_value = int(temp_effect_value)
                                else:
                                    effect_value = str(temp_effect_value)

                        list_of_effects[effect_id] = effect_value
                else:
                    list_of_effects = effects
            else:
                individual_effect = data.split(":")
                if want_list is False:
                    effect_id = individual_effect[0]
                    if effect_id in bscm.convert_to_list:
                        effect_value = individual_effect[1].split(",")
                    else:
                        if individual_effect[1].isnumeric():
                            effect_value = int(individual_effect[1])
                        else:
                            temp_effect_value = individual_effect[1]
                            if temp_effect_value[0] == "-":
                                effect_value = int(temp_effect_value)
                            else:
                                effect_value = str(temp_effect_value)
                    list_of_effects[effect_id] = effect_value
                else:
                    list_of_effects = individual_effect
        else:
            # Data is empty so don't do anything
            pass

        return list_of_effects

    def create_json_datafile(self, excel_workbook="broken_shield_gamedata.xlsx"):
        # We need this for the JSON writing
        bscm = BSCMConfig()

        valid_categories = bscm.node_map_categories

        category_set = set(valid_categories)

        lc = LiveCharacterModel(char_id=999999, live_char_id=999999, player_id=999999)
        lc_dict = lc.dict()
        valid_nodes = list(lc_dict.keys())

        utils = BrokenShieldUtilities()
        json_file = "broken_shield_gamedata"
        count = 0

        excel_path = "./gamedata/" + excel_workbook

        # Open up the Excel workbook for processing
        workbook = load_workbook(
            filename=excel_path,
            read_only=True,
            keep_vba=False,
            data_only=True,
            keep_links=False,
        )
        sheet = workbook.active

        # Define the empty dicts to store the data in
        data_file: dict = {}
        mods_data: dict = {}
        effects_list: list = []
        node_category_list = set()
        node_type_list = set()
        tpb: int = 0
        tps: int = 0
        mod_errors: list = []
        inactive_nodes: int = 0

        # Mod types that do not change Talent Point
        tp_no_change = bscm.tp_no_change

        # Mod types that give 2 bonus Talent Points (Traits)
        tp_bonus = bscm.tp_bonus

        # Mods that cost 1 Talent Point
        tp_cost = bscm.tp_cost

        # All Mod types
        all_mods = set(tp_no_change + tp_bonus + tp_cost)

        # Read each row (except the first as that is headings) and process the data
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Ensure there is actually mod entry!
                # 15 = ACTIVE
                if row[15] == 1:  # Only add mods that are indicated as Active = 1
                    count += 1

                    # 1 = MOD_ID
                    # 2 = NAME
                    # 3 = CATEGORY
                    # 4 = TYPE
                    if row[4]:
                        # Discover if there is a Spent TP or Bonus TP value by mod type
                        if row[4] in tp_bonus:
                            tpb = 2
                            tps = 0
                        elif row[4] in tp_cost:
                            tps = 1
                            tpb = 0
                        else:
                            tps = 0
                            tpb = 0

                        # print(f"TPB = {tpb} and TPS = {tps}")

                    # 5 = CHOOSE TEXT
                    if row[5] == 1:
                        choose_text = True
                    else:
                        choose_text = False

                    # 6 = ALLOW MULTIPLE
                    if row[6] == 1:
                        allow_multiple = True
                    else:
                        allow_multiple = False

                    # 7 = PREREQS = LIST [] or null
                    # 8 = PREREQ ANY ALL
                    # Clean Prerequisites
                    if row[7]:
                        prereq_data = self.split_effects(row[7], True)
                    else:
                        prereq_data = row[7]

                    if row[8] == "all":
                        prereq_any = False
                    else:  # Blank default to Any/True
                        prereq_any = True

                    # 9 = RESTRICTION = LIST [] or null
                    # 10 = RESTRICTION ANY ALL
                    # Clean Restrictions
                    if row[9]:
                        restrict_data = self.split_effects(row[9], True)
                    else:
                        restrict_data = row[9]

                    if row[10] == "all":
                        restrict_any = False
                    else:  # Blank default to Any/True
                        restrict_any = True

                    # 11 = TOUCHED SKILLS = LIST []
                    # Clean Touched Skills: grab list from row 12
                    if row[12]:
                        touch_data = []
                        raw_touch_data = self.split_effects(row[12])
                        new_touch_key = list(raw_touch_data.keys())
                        for y in new_touch_key:
                            if y in valid_nodes:
                                touch_data.append(y.lower())
                            else:
                                touch_data.append(y.lower())
                                logging.error("INVALID NODE FOUND: " + str(y))
                                mod_errors.append(y)
                        # Add in Spent TP or Bonus TP value
                        if tps > 0:
                            touch_data.append("tp_spent")
                        if tpb > 0:
                            touch_data.append("tp_bonus")
                    else:
                        touch_data = row[12]

                    # 12 = EFFECTS = DICT {}
                    # Clean Effects
                    if row[12]:
                        effect_data = self.split_effects(row[12])
                        # Add in Spent TP or Bonus TP value
                        if tps > 0:
                            effect_data["tp_spent"] = tps
                        if tpb > 0:
                            effect_data["tp_bonus"] = tpb
                        if effect_data:
                            new_key = list(effect_data.keys())
                            for x in new_key:
                                if x in effects_list:
                                    pass
                                else:
                                    effects_list.append(x)
                            # Correct "true"/"false" effects for actual true/false
                            #for tf in effect_data:
                            #    if effect_data[tf] == "true":
                            #        effect_data[tf] = True
                            #    elif effect_data[tf] == "false":
                            #        effect_data[tf] = False
                        else:
                            logging.error("Effect data: Null")
                    else:
                        effect_data = row[12]

                    # 13 = BOOK
                    # 14 = PAGE
                    # REF = DICT {}
                    page_ref = (row[13], row[14])
                    if all(page_ref):
                        reference = {str(row[13]): row[14]}
                    else:
                        reference = None

                    # JSON name of Mod
                    mod_id = row[0]
                    mod_details = {
                        "name": row[1],
                        "description": row[2],
                        "category": row[3],
                        "type": row[4],
                        "choose_text": choose_text,
                        "allow_multiple": allow_multiple,
                        "prereqs": prereq_data,
                        "prereq_any": prereq_any,
                        "restriction": restrict_data,
                        "restrict_any": restrict_any,
                        "skills_touched": touch_data,
                        "effects": effect_data,
                        "ref": reference,
                    }
                    mods_data[mod_id] = mod_details

                    # Write to Gamedata DB
                    # [mod_id, name, description, category, mod_type, choose_text, allow_multiple, prereqs,
                    #  prereq_any, restriction, restriction_any, skills_touched, effects, ref]
                    logging.debug("SQLite: writing mod_id " + str(mod_id))

                    self.write_gamedata_db(
                        mod_id,
                        mod_details["name"],
                        mod_details["description"],
                        mod_details["category"],
                        mod_details["type"],
                        mod_details["choose_text"],
                        mod_details["allow_multiple"],
                        json.dumps(mod_details["prereqs"]),
                        mod_details["prereq_any"],
                        json.dumps(mod_details["restriction"]),
                        mod_details["restrict_any"],
                        json.dumps(mod_details["skills_touched"]),
                        json.dumps(mod_details["effects"]),
                        json.dumps(mod_details["ref"]),
                    )

                    logging.debug(
                        str(count)
                        + ": "
                        + str(mod_id)
                        + " ("
                        + str(mod_details["name"])
                        + ")"
                    )
                    node_category_list.add(mod_details["category"])
                    node_type_list.add(mod_details["type"])

                else:
                    inactive_nodes += 1

        # Encase the dict in the overall type: mods_data
        data_file[json_file] = mods_data

        # Write the JSON file
        utils.write_json(data_file, json_file, "w", "./gamedata")

        effects_list.sort()

        # Compare list of categories of previously known mods to check there aren't any new ones / typos

        check_category_sets = node_category_list ^ category_set
        # ^ searches for "not in intersection" of each set

        if len(check_category_sets) > 0:
            process_string0 = (
                "\n--> New/unused mod CATEGORIES encountered in character_dataclasses: \n"
                + str(check_category_sets)
            )
        else:
            process_string0 = "\n--> No new/unused mod CATEGORIES encountered in character_dataclasses."

        # Compare list of previously known mod types to mod types found in gamedata export. We need to know if any
        # new ones have been found
        check_mod_sets = (
            node_type_list ^ all_mods
        )  # ^ searches for "not in intersection" of each set

        if len(check_mod_sets) > 0:
            process_string1 = (
                "\n--> New/unused mod TYPES encountered in character_dataclasses: \n"
                + str(check_mod_sets)
            )
        else:
            process_string1 = (
                "\n--> No new/unused mod TYPES encountered in character_dataclasses."
            )

        logging.warning(
            "\n....................\n"
            + "\nDATA VALIDATION:\n"
            + "\n1. MOD CATEGORIES\nAll Mod Categories encountered during gamedata export ("
            + str(len(node_category_list))
            + " Mod CATEGORIES):\n--> "
            + str(node_category_list)
            + process_string0
            + "\n\n2. MOD TYPES\nAll Mod TYPES encountered during gamedata export ("
            + str(len(node_type_list))
            + " Mod TYPES):\n--> "
            + str(node_type_list)
            + process_string1
            + "\n\n3. MOD ERRORS\nThe following mod ERRORS were noted:\n"
            + str(mod_errors)
            + "\n\n4. INACTIVE MODS: "
            + str(inactive_nodes)
            + " (Active column not set to 1)"
            + "\n\nEND DATA VALIDATION"
            + "\n....................\n"
        )

    @staticmethod
    def create_gamedata_db(
        db="broken_shield_gamedata.sqlite",
        gamedata="gamedata",
        db_path="./gamedata/",
    ):
        # IF broken_shield_gamedata DOESN'T EXIST CREATE IT.
        conn = sqlite3.connect(db_path + db)
        cursor = conn.cursor()
        logging.info(f"SQLite: connecting to db = {db_path}{db}")

        # IF gamedata TABLE DOESN'T EXIST, CREATE IT.  IF IT DOES EXIST DELETE AND RECREATE!
        drop_table_sql = f"DROP TABLE IF EXISTS {gamedata}"

        logging.info(f"SQLite: drop table if exists = {gamedata}")

        create_table_sql = (
            f"CREATE TABLE IF NOT EXISTS {gamedata} (mod_id VARCHAR PRIMARY_KEY NOT NULL, name TEXT, "
            f"description TEXT, category TEXT, type VARCHAR, choose_text BOOL, allow_multiple BOOL, "
            f"prereqs TEXT, prereq_any BOOL, restriction TEXT, restriction_any BOOL, skills_touched TEXT, "
            f"effects TEXT, ref TEXT)"
        )

        logging.info(f"SQLite: create table = {create_table_sql}")
        cursor.execute(drop_table_sql)
        cursor.execute(create_table_sql)
        conn.commit()
        conn.close()

    @staticmethod
    def write_gamedata_db(
        mod_id,
        name,
        description,
        category,
        mod_type,
        choose_text,
        allow_multiple,
        prereqs,
        prereq_any,
        restriction,
        restriction_any,
        skills_touched,
        effects,
        ref,
        db="broken_shield_gamedata.sqlite",
        gamedata="gamedata",
        db_path="./gamedata/",
    ):
        conn = sqlite3.connect(db_path + db)
        cursor = conn.cursor()

        write_sql = (
            f"INSERT INTO {gamedata} (mod_id, name, description, category, type, choose_text, allow_multiple, "
            f"prereqs, prereq_any, restriction, restriction_any, skills_touched, effects, ref) "
            f"VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
        )

        cursor.execute(
            write_sql,
            [
                mod_id,
                name,
                description,
                category,
                mod_type,
                choose_text,
                allow_multiple,
                prereqs,
                prereq_any,
                restriction,
                restriction_any,
                skills_touched,
                effects,
                ref,
            ],
        )

        logging.info(f"SQLite: writing DB entry for {mod_id}")
        conn.commit()
        conn.close()


run_import = ExcelImport()
# Import Data
run_import.create_json_datafile("broken_shield_gamedata.xlsx")
